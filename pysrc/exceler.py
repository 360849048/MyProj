import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Fill, Protection
from copy import copy


def copyCell(src_cell, dst_cell, style=True):
    '''复制一个单元格，不能是合并的单元格，style=False时只复制单元格值，忽略样式'''
    if type(src_cell) != openpyxl.cell.cell.Cell or type(dst_cell) != openpyxl.cell.cell.Cell:
        print("Fatal: copyCell--传入的不是单元格对象")
        return -1

    dst_cell.value = src_cell.value
    if style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.protection = copy(src_cell.protection)
    return 0


def copyCellRange(src_cell_start, src_cell_end, dst_cell_start, style=True):
    '''复制一片单元格选区到指定位置'''
    src_ws = src_cell_start.parent
    dst_ws = dst_cell_start.parent
    if src_ws != src_cell_end.parent:
        print("Fatal: copyCellRange--复制的单元格区域不在同一个sheet中")
        return -1
    if src_cell_start.row > src_cell_end.row or src_cell_start.col_idx > src_cell_end.col_idx:
        print("Fatal: copyCellRange--无法构成复制的单元格区域")
        return -2
    dst_areas_to_merge = []
    for merged_cell in src_ws.merged_cells.ranges:
        merged_cell_start_col = merged_cell.bounds[0]
        merged_cell_start_row = merged_cell.bounds[1]
        merged_cell_end_col = merged_cell.bounds[2]
        merged_cell_end_row = merged_cell.bounds[3]
        areas_pos = _compareAreas(src_cell_start.col_idx, src_cell_start.row, src_cell_end.col_idx, src_cell_end.row,
                                  merged_cell_start_col, merged_cell_start_row, merged_cell_end_col, merged_cell_end_row)
        if areas_pos < 0:
            print("Fatal: copyCellRange--复制的单元格区域中包含了不完整的合并单元格，无法复制，返回码：", areas_pos)
            print(merged_cell)
            return -3
        if areas_pos == 0 or areas_pos == 2:
            dst_areas_to_merge.append([merged_cell_start_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                       merged_cell_start_row - src_cell_start.row + dst_cell_start.row,
                                       merged_cell_end_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                       merged_cell_end_row - src_cell_start.row + dst_cell_start.row])
    for row in range(src_cell_start.row, src_cell_end.row + 1):
        for col in range(src_cell_start.col_idx, src_cell_end.col_idx + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row - src_cell_start.row + dst_cell_start.row,
                                   column=col - src_cell_start.col_idx + dst_cell_start.col_idx)
            if copyCell(src_cell, dst_cell, style) != 0:
                print("Fatal: copyCellRange--单元格复制失败")
                return -4
    for dst_area_to_merge in dst_areas_to_merge:
        dst_ws.merge_cells(start_column=dst_area_to_merge[0],
                           start_row=dst_area_to_merge[1],
                           end_column=dst_area_to_merge[2],
                           end_row=dst_area_to_merge[3])
    return 0


def _compareAreas(area1_start_col, area1_start_row, area1_end_col, area1_end_row,
                  area2_start_col, area2_start_row, area2_end_col, area2_end_row):
    '''比较两处矩形单元格选区的位置关系'''
    if area1_start_col == area2_start_col and \
            area1_start_row == area2_start_row and \
            area1_end_col == area2_end_col and \
            area1_end_row == area2_end_row:
        # 两块区域完全重合
        return 0
    if area1_start_col <= area2_start_col and \
            area1_start_row <= area2_start_row and \
            area1_end_col >= area2_end_col and \
            area1_end_row >= area2_end_row:
        # 区域1包括区域2
        return 2
    if area1_start_col >= area2_start_col and \
            area1_start_row >= area2_start_row and \
            area1_end_col <= area2_end_col and \
            area1_end_row <= area2_end_row:
        # 区域2包括区域1
        return -2
    if (area1_end_col < area2_start_col or area1_end_row < area2_start_row) or \
            (area2_end_col < area1_start_col or area2_end_row < area1_start_row):
        # 两块区域没有任何重合
        return 1
    # 两块区域有部分重合但并不互相包含
    return -1


def searchCells(worksheet, keywords):
    '''根据关键词返回唯一的Cell对象，注意：只能，必须匹配到唯一的单元格,否则返回None'''
    result = []
    if type(worksheet) != openpyxl.worksheet.worksheet.Worksheet:
        print("Fatal: searchCells--不是有效的worksheet对象")
        return None
    str_keywords = str(keywords)
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            if cell.value is not None:
                if str_keywords.upper() in cell.value.upper():
                    result.append(cell)
    if len(result) != 1:
        print("Fatal: searchCells--匹配失败，搜索结果是 ", result)
        return None
    return result[0]


class IOFile:
    def __init__(self, imm_type, main_board_modules=None, board_1_modules=None, big=False):
        ''' imm_type:           'ZEs', 'ZE', 'VE2', 'VE2s'
            main_board_modules: [['CTO163', {'O3': ['中文名', 'EnglishName'], ...}], ['CDM163', {'I5': ['中文名', 'EnglishName'], ...}], ...]
            board_1_modules:    [['CTO163', {'O3': ['中文名', 'EnglishName'], ...}], ['CDM163', {'I5': ['中文名', 'EnglishName'], ...}], ...]
        '''
        self.imm_type = imm_type
        self.main_board_modules = None
        self.main_board_modules_ioconfig = None
        self.board1_1_modules = None
        self.board1_1_modules_ioconfig = None
        self.big = big
        self.src_path = r'./libfiles'
        self.std_filename = None
        self.modulelib_filename = os.path.join(self.src_path, '常用硬件表.xlsx')
        self.std_workbook = None
        self.modulelib_workbook = None

        if main_board_modules is not None:
            self.main_board_modules = []
            self.main_board_modules_ioconfig = []
            for module in main_board_modules:
                self.main_board_modules.append(module[0])
                self.main_board_modules_ioconfig.append(module[1])
        if board_1_modules is not None:
            self.board1_1_modules = []
            self.board1_1_modules_ioconfig = []
            for module in board_1_modules:
                self.board1_1_modules.append(module[0])
                self.board1_1_modules_ioconfig.append(module[1])

        for filename in os.listdir(self.src_path):
            if self.big:
                if filename.upper().startswith('大机 - ' + self.imm_type.upper() + '标准'):
                    if self.std_filename is None:
                        self.std_filename = os.path.join(self.src_path, filename)
                    else:
                        print("Fatal: 匹配到多份标准IO文件：%s  %s" % (self.std_filename, filename))
                        self.std_filename = None
            else:
                if filename.upper().startswith(self.imm_type.upper() + '标准'):
                    if self.std_filename is None:
                        self.std_filename = os.path.join(self.src_path, filename)
                    else:
                        print("Fatal: 匹配到多份标准IO文件：%s  %s" % (self.std_filename, filename))
                        self.std_filename = None

        if self.std_filename is None:
            print("Fatal: 未找到相应标准IO文件，请检查目录")
        else:
            try:
                self.std_workbook = openpyxl.load_workbook(self.std_filename)
            except:
                self.std_workbook = None
                print("Fatal: 试图打开标准IO文件，但它不是正确的xlsx文件")

        try:
            self.modulelib_workbook = openpyxl.load_workbook(self.modulelib_filename)
        except:
            self.modulelib_filename = None
            print("Fatal: 试图打开常用硬件表，但这不是正确的xlsx文件")

    def _copyAndModifyCell(self, src_cell, dst_cell, ioconfiguration=None, style=True):
        '''复制并修改单元格内容，只能修改是 '/空' 或 '/Empty' 单元格
            ioconfiguration: {'I2': xxx, 'O4': xxx, 'AI1': xxx, 'AO2': xxx, 'TI1': xxx, 'TO1': xxx}
        '''
        if type(src_cell) != openpyxl.cell.cell.Cell or type(dst_cell) != openpyxl.cell.cell.Cell:
            print("Fatal: copyCell--传入的不是单元格对象")
            return -1
        src_value = src_cell.value
        if str(src_value).startswith('/空') or str(src_value).startswith('/Empty'):
            if ioconfiguration is not None:
                for key, value in ioconfiguration.items():
                    if '/空' + key.upper() == str(src_value):
                        src_value = str(value[0])
                        dst_cell.value = src_value
                    if '/Empty' + key.upper() == str(src_value):
                        src_value = str(value[1])
                        dst_cell.value = src_value
        else:
            dst_cell.value = src_value
        if style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.protection = copy(src_cell.protection)
        return 0

    def _copyAndModifyCellRange(self, src_cell_start, src_cell_end, dst_cell_start, ioconfiguration=None, style=True):
        '''复制并修改一片单元格选区到指定位置
            ioconfiguration: {'I2': xxx, 'O4': xxx, 'AI1': xxx, 'AO2': xxx}
        '''
        src_ws = src_cell_start.parent
        dst_ws = dst_cell_start.parent
        if src_ws != src_cell_end.parent:
            print("Fatal: _copyAndModifyCellRange--复制的单元格区域不在同一个sheet中")
            return -1
        if src_cell_start.row > src_cell_end.row or src_cell_start.col_idx > src_cell_end.col_idx:
            print("Fatal: _copyAndModifyCellRange--无法构成复制的单元格区域")
            return -2
        dst_areas_to_merge = []
        for merged_cell in src_ws.merged_cells.ranges:
            merged_cell_start_col = merged_cell.bounds[0]
            merged_cell_start_row = merged_cell.bounds[1]
            merged_cell_end_col = merged_cell.bounds[2]
            merged_cell_end_row = merged_cell.bounds[3]
            areas_pos = _compareAreas(src_cell_start.col_idx, src_cell_start.row, src_cell_end.col_idx,
                                      src_cell_end.row,
                                      merged_cell_start_col, merged_cell_start_row, merged_cell_end_col,
                                      merged_cell_end_row)
            if areas_pos < 0:
                print("Fatal: _copyAndModifyCellRange--复制的单元格区域中包含了不完整的合并单元格，无法复制，返回码：", areas_pos)
                print(merged_cell)
                return -3
            if areas_pos == 0 or areas_pos == 2:
                dst_areas_to_merge.append([merged_cell_start_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                           merged_cell_start_row - src_cell_start.row + dst_cell_start.row,
                                           merged_cell_end_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                           merged_cell_end_row - src_cell_start.row + dst_cell_start.row])
        for row in range(src_cell_start.row, src_cell_end.row + 1):
            for col in range(src_cell_start.col_idx, src_cell_end.col_idx + 1):
                src_cell = src_ws.cell(row=row, column=col)
                dst_cell = dst_ws.cell(row=row - src_cell_start.row + dst_cell_start.row,
                                       column=col - src_cell_start.col_idx + dst_cell_start.col_idx)
                if self._copyAndModifyCell(src_cell, dst_cell, ioconfiguration=ioconfiguration, style=style) != 0:
                    print("Fatal: _copyAndModifyCellRange--单元格复制失败")
                    return -4
        for dst_area_to_merge in dst_areas_to_merge:
            dst_ws.merge_cells(start_column=dst_area_to_merge[0],
                               start_row=dst_area_to_merge[1],
                               end_column=dst_area_to_merge[2],
                               end_row=dst_area_to_merge[3])
        return 0

    def copyMainBoardModulesInfo(self):
        ''' 往 硬件配置 中复制主底板扩展模块信息 '''
        if self.main_board_modules is None:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 在  '硬件配置' 表中定位 '主底板扩展槽' 单元格
        result = searchCells(hardware_info_worksheet, '主底板扩展槽')
        if result is None:
            print("Fatal: 在 %s->%s 中匹配关键字: '主底板扩展槽' 失败" % (self.std_filename, hardware_info_worksheet.title))
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None or first_cell.value == '扩展底板一':
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None and first_cell.value != '扩展底板一':
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        # 向 '硬件配置' 表复制主底板扩展槽模块信息
        if self.big:
            last_module = 'CIO021'
        else:
            last_module = None
        for module in self.main_board_modules:
            if last_module is not None and module.upper() == last_module.upper():
                module_num = int(
                    hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3).value) + 1
                hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3,
                                             value='+' + str(module_num))
                continue
            hardware_info_worksheet.insert_rows(cur_work_row)
            src_end_cell = searchCells(module_content_worksheet, str(module))
            if src_end_cell is None:
                print("Fatal: 在 %s->%s 中匹配关键字'%s'失败 " % (self.modulelib_filename, module_content_worksheet.title, module))
                return -2
            src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
            dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
                print("Fatal: 单元格区域复制失败")
                return -3
            hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')
            cur_work_row += 1
            last_module = module
        return 0

    def copyMainBoardModules(self):
        ''' 往 主底板扩展槽 复制模块信息 '''
        if self.main_board_modules is None:
            return 0
        main_board_hardware_worksheet = self.std_workbook['主底板扩展槽IO']
        cur_work_row = main_board_hardware_worksheet.max_row + 1
        cur_work_column = 1
        # 获取各模块对应的Worksheet对象
        modulelib_worksheets = []
        for module in self.main_board_modules:
            for sheetname in self.modulelib_workbook.sheetnames:
                temp_sheetname = ''.join([a for a in sheetname if a != ' '])
                if module.upper() == temp_sheetname.upper():
                    modulelib_worksheets.append(self.modulelib_workbook[sheetname])
        if len(modulelib_worksheets) != len(self.main_board_modules):
            print("Fatal: 在'%s'中寻找对应模块sheet标签失败" % self.modulelib_workbook)
            return -1
        # 向 '主底板扩展槽' 复制模块
        for i in range(len(self.main_board_modules)):
            src_start_cell = modulelib_worksheets[i].cell(row=2, column=1)
            src_end_cell = modulelib_worksheets[i].cell(row=modulelib_worksheets[i].max_row, column=4)
            dst_start_cell = main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, ioconfiguration=self.main_board_modules_ioconfig[i]) != 0:
                print("Fatal: 单元格区域复制失败")
                return -2
            if self.big:
                if self.imm_type.upper() != 'VE2':
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 1)
                else:
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 8)
            else:
                if self.imm_type.upper() != 'VE2':
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i)
                else:
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 7)
            cur_work_row += modulelib_worksheets[i].max_row - 1
        return 0

    def copyBoard1ModulesInfo(self):
        ''' 往 硬件配置 中复制扩展底板一模块信息 '''
        if self.board1_1_modules is None:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 定位“扩展底板一”单元格
        result = searchCells(hardware_info_worksheet, '扩展底板一')
        if result is None:
            print("Fatal: 在%s中匹配关键字: '扩展底板一' 失败" % self.std_filename)
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None or first_cell.value == '扩展底板二':
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None and first_cell.value != '扩展底板二':
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        last_module = None
        for module in self.board1_1_modules:
            if last_module is not None and module.upper() == last_module.upper():
                module_num = int(
                    hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3).value) + 1
                hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3,
                                             value='+' + str(module_num))
                continue
            hardware_info_worksheet.insert_rows(cur_work_row)
            src_end_cell = searchCells(module_content_worksheet, str(module))
            if src_end_cell is None:
                print("Fatal: 在%s中匹配关键字'%s'失败 " % (self.modulelib_filename, module))
                return -2
            src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
            dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
                print("Fatal: 单元格区域复制失败")
                return -3
            hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')
            cur_work_row += 1
            last_module = module
        return 0

    def copyBoard1Modules(self):
        ''' 往 扩展底板一IO 复制模块信息 '''
        if self.board1_1_modules is None:
            return 0
        board_1_hardware_worksheet = self.std_workbook['扩展底板一IO']
        cur_work_row = board_1_hardware_worksheet.max_row + 1
        cur_work_column = 1
        # 获取各模块对应的Worksheet对象
        modulelib_worksheets = []
        for module in self.board1_1_modules:
            for sheetname in self.modulelib_workbook.sheetnames:
                temp_sheetname = ''.join([a for a in sheetname if a != ' '])
                if module.upper() == temp_sheetname.upper():
                    modulelib_worksheets.append(self.modulelib_workbook[sheetname])
        if len(modulelib_worksheets) != len(self.board1_1_modules):
            print("Fatal: 在'%s'中寻找对应模块sheet标签失败" % self.modulelib_workbook)
            return -1
        # 向 '扩展底板一IO' 复制模块，忽略第一块VARAN模块
        for i in range(1, len(self.board1_1_modules)):
            src_start_cell = modulelib_worksheets[i].cell(row=2, column=1)
            src_end_cell = modulelib_worksheets[i].cell(row=modulelib_worksheets[i].max_row, column=4)
            dst_start_cell = board_1_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, ioconfiguration=self.board1_1_modules_ioconfig[i]) != 0:
                print("Fatal: 单元格区域复制失败")
                return -2
            board_1_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i - 1)
            cur_work_row += modulelib_worksheets[i].max_row - 1
        return 0

    def saveAs(self, filename):
        self.std_workbook.save(filename)
        return 0



if __name__ == '__main__':
    iofile = IOFile(imm_type='ve2s',
                    main_board_modules=[['cdm163', {'I1': ['点1', 'id 1']}], ['cio011', {'O2': ['点45', 'id 45'], 'Ai2': ['模腔压力', 'Mold Pressure']}], ['cto163', {'O1': ['点41', 'ID 41']}]],
                    board_1_modules=[['civ512', {}], ['cai888', {}]],
                    big=True)
    iofile.copyMainBoardModulesInfo()
    iofile.copyMainBoardModules()
    iofile.copyBoard1ModulesInfo()
    iofile.copyBoard1Modules()
    iofile.saveAs('../1.xlsx')
