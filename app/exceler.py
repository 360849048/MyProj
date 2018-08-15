import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Fill, Protection
from copy import copy


def copyCell(src_cell, dst_cell, style=True):
    '''复制一个单元格，不能是合并的单元格，style=False时只复制单元格值，忽略样式'''
    if type(src_cell) != openpyxl.cell.cell.Cell or type(dst_cell) != openpyxl.cell.cell.Cell:
        print("Fatal: copyCell--传入的不是单元格对象")
        return -1

    dst_cell.value = src_cell.value
    if style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.protection = copy(src_cell.protection)
    return 0


def copyCellRange(src_cell_start, src_cell_end, dst_cell_start, style=True):
    '''复制一片单元格选区到指定位置'''
    src_ws = src_cell_start.parent
    dst_ws = dst_cell_start.parent
    if src_ws != src_cell_end.parent:
        print("Fatal: copyCellRange--复制的单元格区域不在同一个sheet中")
        return -1
    if src_cell_start.row > src_cell_end.row or src_cell_start.col_idx > src_cell_end.col_idx:
        print("Fatal: copyCellRange--无法构成复制的单元格区域")
        return -2
    dst_areas_to_merge = []
    for merged_cell in src_ws.merged_cells.ranges:
        merged_cell_start_col = merged_cell.bounds[0]
        merged_cell_start_row = merged_cell.bounds[1]
        merged_cell_end_col = merged_cell.bounds[2]
        merged_cell_end_row = merged_cell.bounds[3]
        areas_pos = _compareAreas(src_cell_start.col_idx, src_cell_start.row, src_cell_end.col_idx, src_cell_end.row,
                                  merged_cell_start_col, merged_cell_start_row, merged_cell_end_col, merged_cell_end_row)
        if areas_pos < 0:
            print("Fatal: copyCellRange--复制的单元格区域中包含了不完整的合并单元格，无法复制，返回码：", areas_pos)
            print(merged_cell)
            return -3
        if areas_pos == 0 or areas_pos == 2:
            dst_areas_to_merge.append([merged_cell_start_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                       merged_cell_start_row - src_cell_start.row + dst_cell_start.row,
                                       merged_cell_end_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                       merged_cell_end_row - src_cell_start.row + dst_cell_start.row])
    for row in range(src_cell_start.row, src_cell_end.row + 1):
        for col in range(src_cell_start.col_idx, src_cell_end.col_idx + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row - src_cell_start.row + dst_cell_start.row,
                                   column=col - src_cell_start.col_idx + dst_cell_start.col_idx)
            if copyCell(src_cell, dst_cell, style) != 0:
                print("Fatal: copyCellRange--单元格复制失败")
                return -4
    for dst_area_to_merge in dst_areas_to_merge:
        dst_ws.merge_cells(start_column=dst_area_to_merge[0],
                           start_row=dst_area_to_merge[1],
                           end_column=dst_area_to_merge[2],
                           end_row=dst_area_to_merge[3])
    return 0


def _compareAreas(area1_start_col, area1_start_row, area1_end_col, area1_end_row,
                  area2_start_col, area2_start_row, area2_end_col, area2_end_row):
    '''比较两处矩形单元格选区的位置关系'''
    if area1_start_col == area2_start_col and \
            area1_start_row == area2_start_row and \
            area1_end_col == area2_end_col and \
            area1_end_row == area2_end_row:
        # 两块区域完全重合
        return 0
    if area1_start_col <= area2_start_col and \
            area1_start_row <= area2_start_row and \
            area1_end_col >= area2_end_col and \
            area1_end_row >= area2_end_row:
        # 区域1包括区域2
        return 2
    if area1_start_col >= area2_start_col and \
            area1_start_row >= area2_start_row and \
            area1_end_col <= area2_end_col and \
            area1_end_row <= area2_end_row:
        # 区域2包括区域1
        return -2
    if (area1_end_col < area2_start_col or area1_end_row < area2_start_row) or \
            (area2_end_col < area1_start_col or area2_end_row < area1_start_row):
        # 两块区域没有任何重合
        return 1
    # 两块区域有部分重合但并不互相包含
    return -1


def searchCells(worksheet, keywords, strict=False):
    '''
    根据搜索词返回唯一的Cell对象
    普通匹配时候，不区分大小写，单元格内容包含keywords即OK
    严格模式下，单元格内容必须与keywords一模一样，大小写敏感，但前后单元格前后的空格不会影响结果
    注意：只能，任何模式工作下必须只匹配到唯一的单元格,否则返回None
    '''
    result = []
    if type(worksheet) != openpyxl.worksheet.worksheet.Worksheet:
        print("Fatal: searchCells--不是有效的worksheet对象")
        return None
    str_keywords = str(keywords)
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            if cell.value is not None:
                if not strict:
                    if str_keywords.upper() in cell.value.upper():
                        result.append(cell)
                else:
                    if str_keywords == cell.value.strip(' '):
                        result.append(cell)
    if len(result) != 1:
        print("Fatal: searchCells--匹配失败，因为匹配到多个单元格，搜索结果是 ", result)
        return None
    return result[0]


class IOFile:
    def __init__(self, imm_type, main_board_modules=None, board_1_modules=None, big=False,
                 evaluation_num=None, production_num=None, type_string=None, customer=None,
                 safety_standard=None, technical_clause=None, dual_inj=False, external_hotrunner_num=0,
                 energy_dee=False, varan_conn_module_pos=0):
        ''' imm_type:           'ZEs', 'ZE', 'VE2', 'VE2s'
            main_board_modules: [['CTO163', {'O3': ['中文名', 'EnglishName'], ...}], ['CDM163', {'I5': ['中文名', 'EnglishName'], ...}], ...]
            board_1_modules:    [['CIV512', {}], ['CTO163', {'O3': ['中文名', 'EnglishName'], ...}], ['CDM163', {'I5': ['中文名', 'EnglishName'], ...}], ...]
            varan_conn_module_pos:  0：如果有Varan连接模块，则配置在KEB之后；   1：如果有Varan连接模块，则配置在KEB之前
        '''
        self.imm_type = imm_type
        self.main_board_modules = None
        self.main_board_modules_ioconfig = None
        self.board1_1_modules = None
        self.board1_1_modules_ioconfig = None
        self.varan_conn_module_pos = int(varan_conn_module_pos)
        self.big = big
        self.src_path = r'./app/libfiles'
        self.std_filename = None
        self.modulelib_filename = os.path.join(self.src_path, '常用硬件表.xlsx')
        self.std_workbook = None
        self.modulelib_workbook = None
        # 待填入表“机器信息”的信息
        self.evaluation_num = evaluation_num
        self.production_num = production_num
        self.type_string = type_string
        self.customer = customer
        self.safety_standard = safety_standard
        self.technical_clause = technical_clause
        # “硬件排布”双注射信息
        self.dual_inj = dual_inj
        # 外置热流道组数
        self.external_hotrunner_num = external_hotrunner_num
        # 能耗模块DEE
        self.energy_dee = energy_dee

        if main_board_modules is not None:
            self.main_board_modules = []
            self.main_board_modules_ioconfig = []
            for module in main_board_modules:
                self.main_board_modules.append(module[0])
                self.main_board_modules_ioconfig.append(module[1])
        if board_1_modules is not None:
            self.board1_1_modules = []
            self.board1_1_modules_ioconfig = []
            for module in board_1_modules:
                self.board1_1_modules.append(module[0])
                self.board1_1_modules_ioconfig.append(module[1])

        for filename in os.listdir(self.src_path):
            if self.big:
                if filename.upper().startswith('大机 - ' + self.imm_type.upper() + '标准'):
                    if self.std_filename is None:
                        self.std_filename = os.path.join(self.src_path, filename)
                    else:
                        print("Fatal: 匹配到多份标准IO文件：%s  %s" % (self.std_filename, filename))
                        self.std_filename = None
            else:
                if filename.upper().startswith(self.imm_type.upper() + '标准'):
                    if self.std_filename is None:
                        self.std_filename = os.path.join(self.src_path, filename)
                    else:
                        print("Fatal: 匹配到多份标准IO文件：%s  %s" % (self.std_filename, filename))
                        self.std_filename = None

        if self.std_filename is None:
            print("Fatal: 未找到相应标准IO文件，请检查目录")
        else:
            try:
                self.std_workbook = openpyxl.load_workbook(self.std_filename)
            except:
                self.std_workbook = None
                print("Fatal: 试图打开标准IO文件，但它不是正确的xlsx文件")

        try:
            self.modulelib_workbook = openpyxl.load_workbook(self.modulelib_filename)
        except:
            self.modulelib_filename = None
            print("Fatal: 试图打开常用硬件表，但这不是正确的xlsx文件")

    def _copyAndModifyCell(self, src_cell, dst_cell, ioconfiguration=None, style=True):
        '''复制并修改单元格内容，只能修改是 '/空' 或 '/Empty' 单元格
            ioconfiguration: {'DI2': (中文名, EnglishName), 'DO4': (中文名, EnglishName), ...}
        '''
        if type(src_cell) != openpyxl.cell.cell.Cell or type(dst_cell) != openpyxl.cell.cell.Cell:
            print("Fatal: copyCell--传入的不是单元格对象")
            return -1
        src_value = src_cell.value
        if str(src_value).startswith('/空') or str(src_value).startswith('/Empty'):
            if ioconfiguration is not None:
                for key, value in ioconfiguration.items():
                    if '/空' + key.upper() == str(src_value):
                        src_value = str(value[0])
                        dst_cell.value = src_value
                    if '/Empty' + key.upper() == str(src_value):
                        src_value = str(value[1])
                        dst_cell.value = src_value
        else:
            dst_cell.value = src_value
        if style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.protection = copy(src_cell.protection)
        return 0

    def _copyAndModifyCellRange(self, src_cell_start, src_cell_end, dst_cell_start, ioconfiguration=None, style=True):
        '''复制并修改一片单元格选区到指定位置
            ioconfiguration: {'DI2': (中文名, EnglishName), 'DO4': (中文名, EnglishName), ...}
        '''
        src_ws = src_cell_start.parent
        dst_ws = dst_cell_start.parent
        if src_ws != src_cell_end.parent:
            print("Fatal: _copyAndModifyCellRange--参数src_cell_start和src_cell_end不在同一个sheet")
            return -1
        if src_cell_start.row > src_cell_end.row or src_cell_start.col_idx > src_cell_end.col_idx:
            print("Fatal: _copyAndModifyCellRange--无法构成复制的单元格区域")
            return -2
        dst_areas_to_merge = []
        for merged_cell in src_ws.merged_cells.ranges:
            merged_cell_start_col = merged_cell.bounds[0]
            merged_cell_start_row = merged_cell.bounds[1]
            merged_cell_end_col = merged_cell.bounds[2]
            merged_cell_end_row = merged_cell.bounds[3]
            areas_pos = _compareAreas(src_cell_start.col_idx, src_cell_start.row, src_cell_end.col_idx,
                                      src_cell_end.row,
                                      merged_cell_start_col, merged_cell_start_row, merged_cell_end_col,
                                      merged_cell_end_row)
            if areas_pos < 0:
                print("Fatal: _copyAndModifyCellRange--复制的单元格区域中包含了不完整的合并单元格，无法复制，返回码：", areas_pos)
                print(merged_cell)
                return -3
            if areas_pos == 0 or areas_pos == 2:
                dst_areas_to_merge.append([merged_cell_start_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                           merged_cell_start_row - src_cell_start.row + dst_cell_start.row,
                                           merged_cell_end_col - src_cell_start.col_idx + dst_cell_start.col_idx,
                                           merged_cell_end_row - src_cell_start.row + dst_cell_start.row])
        for row in range(src_cell_start.row, src_cell_end.row + 1):
            for col in range(src_cell_start.col_idx, src_cell_end.col_idx + 1):
                src_cell = src_ws.cell(row=row, column=col)
                dst_cell = dst_ws.cell(row=row - src_cell_start.row + dst_cell_start.row,
                                       column=col - src_cell_start.col_idx + dst_cell_start.col_idx)
                if self._copyAndModifyCell(src_cell, dst_cell, ioconfiguration=ioconfiguration, style=style) != 0:
                    print("Fatal: _copyAndModifyCellRange--单元格复制失败")
                    return -4
        for dst_area_to_merge in dst_areas_to_merge:
            dst_ws.merge_cells(start_column=dst_area_to_merge[0],
                               start_row=dst_area_to_merge[1],
                               end_column=dst_area_to_merge[2],
                               end_row=dst_area_to_merge[3])
        return 0

    def copyMainBoardModulesInfo(self):
        ''' 往 硬件配置 中复制主底板扩展模块信息 '''
        if self.main_board_modules is None:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 在  '硬件配置' 表中定位 '主底板扩展槽' 单元格
        result = searchCells(hardware_info_worksheet, '主底板扩展槽')
        if result is None:
            print("Fatal: 在 %s->%s 中匹配关键字: '主底板扩展槽' 失败" % (self.std_filename, hardware_info_worksheet.title))
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None or first_cell.value == '扩展底板一':
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None and first_cell.value != '扩展底板一':
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        # 向 '硬件配置' 表复制主底板扩展槽模块信息
        if self.big:
            last_module = 'CIO021'
        else:
            last_module = None
        for module in self.main_board_modules:
            if last_module is not None and module.upper() == last_module.upper():
                module_num = int(
                    hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3).value) + 1
                hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3,
                                             value='+' + str(module_num))
                continue
            hardware_info_worksheet.insert_rows(cur_work_row)
            src_end_cell = searchCells(module_content_worksheet, str(module))
            if src_end_cell is None:
                print("Fatal: 在 %s->%s 中匹配关键字'%s'失败 " % (self.modulelib_filename, module_content_worksheet.title, module))
                return -2
            src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
            dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
                print("Fatal: 单元格区域复制失败")
                return -3
            hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')
            cur_work_row += 1
            last_module = module
        return 0

    def copyMainBoardModules(self):
        ''' 往 主底板扩展槽 复制模块信息 '''
        if self.main_board_modules is None:
            return 0
        main_board_hardware_worksheet = self.std_workbook['主底板扩展槽IO']
        cur_work_row = main_board_hardware_worksheet.max_row + 1
        cur_work_column = 1
        # 获取各模块对应的Worksheet对象
        modulelib_worksheets = []
        for module in self.main_board_modules:
            for sheetname in self.modulelib_workbook.sheetnames:
                temp_sheetname = ''.join([a for a in sheetname if a != ' '])
                if module.upper() == temp_sheetname.upper():
                    modulelib_worksheets.append(self.modulelib_workbook[sheetname])
        if len(modulelib_worksheets) != len(self.main_board_modules):
            print("Fatal: 在'%s'中寻找对应模块sheet标签失败" % self.modulelib_workbook)
            return -1
        # 向 '主底板扩展槽' 复制模块
        for i in range(len(self.main_board_modules)):
            src_start_cell = modulelib_worksheets[i].cell(row=2, column=1)
            src_end_cell = modulelib_worksheets[i].cell(row=modulelib_worksheets[i].max_row, column=4)
            dst_start_cell = main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, ioconfiguration=self.main_board_modules_ioconfig[i]) != 0:
                print("Fatal: 单元格区域复制失败")
                return -2
            if self.big:
                if self.imm_type.upper() != 'VE2':
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 1)
                else:
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 8)
            else:
                if self.imm_type.upper() != 'VE2':
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i)
                else:
                    main_board_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i + 7)
            cur_work_row += modulelib_worksheets[i].max_row - 1
        return 0

    def copyBoard1ModulesInfo(self):
        ''' 往 硬件配置 中复制扩展底板一模块信息 '''
        if self.board1_1_modules is None:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 定位“扩展底板一”单元格
        result = searchCells(hardware_info_worksheet, '扩展底板一')
        if result is None:
            print("Fatal: 在%s中匹配关键字: '扩展底板一' 失败" % self.std_filename)
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None or first_cell.value == '扩展底板二':
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None and first_cell.value != '扩展底板二':
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        last_module = None
        for module in self.board1_1_modules:
            if last_module is not None and module.upper() == last_module.upper():
                module_num = int(
                    hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3).value) + 1
                hardware_info_worksheet.cell(row=cur_work_row - 1, column=cur_work_column + 3,
                                             value='+' + str(module_num))
                continue
            hardware_info_worksheet.insert_rows(cur_work_row)
            src_end_cell = searchCells(module_content_worksheet, str(module))
            if src_end_cell is None:
                print("Fatal: 在%s中匹配关键字'%s'失败 " % (self.modulelib_filename, module))
                return -2
            src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
            dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
                print("Fatal: 单元格区域复制失败")
                return -3
            hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')
            cur_work_row += 1
            last_module = module
        return 0

    def copyBoard1Modules(self):
        ''' 往 扩展底板一IO 复制模块信息 '''
        if self.board1_1_modules is None:
            return 0
        board_1_hardware_worksheet = self.std_workbook['扩展底板一IO']
        cur_work_row = board_1_hardware_worksheet.max_row + 1
        cur_work_column = 1
        # 获取各模块对应的Worksheet对象
        modulelib_worksheets = []
        for module in self.board1_1_modules:
            for sheetname in self.modulelib_workbook.sheetnames:
                temp_sheetname = ''.join([a for a in sheetname if a != ' '])
                if module.upper() == temp_sheetname.upper():
                    modulelib_worksheets.append(self.modulelib_workbook[sheetname])
        if len(modulelib_worksheets) != len(self.board1_1_modules):
            print("Fatal: 在'%s'中寻找对应模块sheet标签失败" % self.modulelib_workbook)
            return -1
        # 向 '扩展底板一IO' 复制模块，忽略第一块VARAN模块
        for i in range(1, len(self.board1_1_modules)):
            src_start_cell = modulelib_worksheets[i].cell(row=2, column=1)
            src_end_cell = modulelib_worksheets[i].cell(row=modulelib_worksheets[i].max_row, column=4)
            dst_start_cell = board_1_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, ioconfiguration=self.board1_1_modules_ioconfig[i]) != 0:
                print("Fatal: 单元格区域复制失败")
                return -2
            board_1_hardware_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i - 1)
            cur_work_row += modulelib_worksheets[i].max_row - 1
        return 0

    def modifyImmInfo(self):
        ''' 填写机器信息以及修改注射keb2信息和DEE信息和Varan连接模块信息 '''
        # 修改表'机器信息'内容
        imm_info_worksheet = self.std_workbook['机器信息']
        def _modify(title, value):
            if value is not None:
                result = searchCells(imm_info_worksheet, title)
                if result is None:
                    print("Fatal: 在 %s->%s 中匹配关键字: %s 失败" % (self.std_filename, imm_info_worksheet.title, title))
                else:
                    imm_info_worksheet.cell(row=result.row + 1, column=result.col_idx, value=value)
        _modify('合同订单编号', self.evaluation_num)
        _modify('客户', self.customer)
        _modify('机型', self.type_string)
        _modify('生产订单号', self.production_num)
        _modify('安全标准', self.safety_standard)
        _modify('软件相关条款', self.technical_clause)

        # 修改 注射keb2
        if self.dual_inj:
            hardware_sort = self.std_workbook['硬件排布']
            keb1_cell = searchCells(hardware_sort, '注射KEB1')
            if keb1_cell is not None:
                keb2_cell = hardware_sort.cell(row=keb1_cell.row, column=keb1_cell.col_idx + 1)
                copyCell(keb1_cell, keb2_cell)
                keb2_cell.value = '注射KEB2'
            if self.board1_1_modules is not None and len(self.board1_1_modules) >= 1 and self.varan_conn_module_pos == 0:
                # 在注射KEB2之后添加CIV
                varan_conn_module_cell = hardware_sort.cell(row=keb1_cell.row, column=keb1_cell.col_idx + 2)
                copyCell(keb1_cell, varan_conn_module_cell)
                varan_conn_module_cell.value = self.board1_1_modules[0].upper()
        else:
            if self.board1_1_modules is not None and len(self.board1_1_modules) >= 1 and self.varan_conn_module_pos == 0:
                # 在注射KEB1之后添加CIV
                hardware_sort = self.std_workbook['硬件排布']
                keb1_cell = searchCells(hardware_sort, '注射KEB1')
                if keb1_cell is not None:
                    varan_conn_module_cell = hardware_sort.cell(row=keb1_cell.row, column=keb1_cell.col_idx + 1)
                    copyCell(keb1_cell, varan_conn_module_cell)
                    varan_conn_module_cell.value = self.board1_1_modules[0].upper()
        # 在KEB之前添加CIV（self.varan_conn_module_pos == 1）
        if self.board1_1_modules is not None and len(self.board1_1_modules) >= 1 and self.varan_conn_module_pos == 1:
            hardware_sort = self.std_workbook['硬件排布']
            if self.imm_type.upper() == 'VE2' or self.imm_type.upper() == 'VE2S':
                ejekeb_cell = searchCells(hardware_sort, '顶出KEB')
            else:
                ejekeb_cell = searchCells(hardware_sort, '液压伺服KEB1')
            if ejekeb_cell is not None:
                cur_work_col = ejekeb_cell.col_idx
                hardware_sort.insert_cols(cur_work_col)
                varan_conn_module_cell = hardware_sort.cell(row=ejekeb_cell.row, column=cur_work_col)
                copyCell(ejekeb_cell, varan_conn_module_cell)
                varan_conn_module_cell.value = self.board1_1_modules[0].upper()
        # 修改 DEE021，插入到顶出KEB(VE)/液压伺服KEB1(ZE)之前
        if self.energy_dee:
            hardware_sort = self.std_workbook['硬件排布']
            if self.imm_type.upper() == 'VE2' or self.imm_type.upper() == 'VE2S':
                ejekeb_cell = searchCells(hardware_sort, '顶出KEB')
            else:
                ejekeb_cell = searchCells(hardware_sort, '液压伺服KEB1')
            if ejekeb_cell is not None:
                cur_work_col = ejekeb_cell.col_idx
                hardware_sort.insert_cols(cur_work_col)
                dee_cell = hardware_sort.cell(row=ejekeb_cell.row, column=cur_work_col)
                copyCell(ejekeb_cell, dee_cell)
                dee_cell.value = 'DEE021'

    def modifyDefaultIO(self, io_type, origin_io_name, new_io_cname, new_io_ename):
        ''' 修改主底板或主底板扩展槽(VE2)上默认的IO
            io_type: IO的类型，可以是DI, DO, AI, AO, TI, TO
            origin_io_name: -默认IO的中文名，为提高定位精准度，IO的中文名必须与单元格内容全词匹配
            new_io_cname: 替换后的IO中文名
            new_io_ename: 替换后的IO英文名
        '''
        if self.imm_type.upper() == 'VE2':
            default_io_ws = self.std_workbook['主底板扩展槽IO']
        else:
            default_io_ws = self.std_workbook['主底板IO']
        # 定位到需要替换的IO点(中文)
        target_cio_cell = searchCells(default_io_ws, origin_io_name, strict=True)
        if target_cio_cell is None:
            print("Fatal: modifyDefaultIO修改主底板默认IO时，无法定位到具体的IO点")
            return -1
        # 验证io类型是否一致
        act_io_type = default_io_ws.cell(row=target_cio_cell.row, column=target_cio_cell.col_idx - 1).value
        act_1st_letter = act_io_type.upper()[0]
        if act_1st_letter != io_type.upper()[0]:
            if act_1st_letter == 'I' and io_type.upper() == 'DI':
                pass
            elif act_1st_letter == 'O' and io_type.upper() == 'DO':
                pass
            else:
                print("Fatal: modifyDefaultIO搜索到的IO类型与要求不符，无法定位到具体的IO点")
                return -2
        target_cio_cell.value = new_io_cname
        # 修改IO点的英文名
        default_io_ws.cell(row=target_cio_cell.row, column=target_cio_cell.col_idx + 1, value=new_io_ename)
        # 从别的地方复制样式(红色背景)后修改 '更改' 列的内容('√')
        modify_check_cell = default_io_ws.cell(row=target_cio_cell.row, column=target_cio_cell.col_idx + 2)
        src_style_cell = searchCells(default_io_ws, '更改', strict=True)
        copyCell(src_cell=src_style_cell, dst_cell=modify_check_cell, style=True)
        modify_check_cell.value = '√'
        return 0

    def copyExternalHotrunnerInfo(self):
        ''' 往 硬件配置 中复制外置热流道信息 '''
        if self.external_hotrunner_num == 0:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 定位“热流道箱”单元格
        result = searchCells(hardware_info_worksheet, '热流道箱')
        if result is None:
            print("Fatal: 在%s中匹配关键字: '热流道箱' 失败" % self.std_filename)
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None or first_cell.value == '特殊功能模块':
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None and first_cell.value != '特殊功能模块':
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        # 复制CCP521模块信息
        hardware_info_worksheet.insert_rows(cur_work_row)
        src_end_cell = searchCells(module_content_worksheet, 'CCP521')
        if src_end_cell is None:
            print("Fatal: 在 %s->%s 中匹配关键字'CCP521'失败 " % (self.modulelib_filename, module_content_worksheet.title))
            return -2
        src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
        dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
            print("Fatal: 单元格区域复制失败")
            return -3
        hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')
        cur_work_row += 1
        # 复制CAI888模块信息
        hardware_info_worksheet.insert_rows(cur_work_row)
        src_end_cell = searchCells(module_content_worksheet, 'CAI888')
        if src_end_cell is None:
            print("Fatal: 在 %s->%s 中匹配关键字'CAI888'失败 " % (self.modulelib_filename, module_content_worksheet.title))
            return -2
        src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
        dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
            print("Fatal: 单元格区域复制失败")
            return -3
        hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+' + str(self.external_hotrunner_num))
        return 0

    def copyExternalHotrunnerModule(self):
        ''' 往 热流道箱IO 复制CAI888模块 '''
        if self.external_hotrunner_num == 0:
            return 0
        hotrunner_io_worksheet = self.std_workbook['热流道箱IO']
        cur_work_row = hotrunner_io_worksheet.max_row + 1
        cur_work_column = 1
        # 获取CAI888模块所在的Worksheet对象
        cai888_worksheets = None
        for sheetname in self.modulelib_workbook.sheetnames:
            temp_sheetname = ''.join([a for a in sheetname if a != ' '])
            if 'CAI888' == temp_sheetname.upper():
                if cai888_worksheets is None:
                    cai888_worksheets = self.modulelib_workbook[sheetname]
                else:
                    print("Fatal: 在'%s'中找到了多个匹配CAI888的标签" % self.modulelib_workbook)
                    return -1
        if cai888_worksheets is None:
            print("Fatal: 在'%s'中寻找CAI888模块对应标签失败" % self.modulelib_workbook)
            return -1
        # 向‘热流道箱IO’复制CAI888模块
        src_start_cell = cai888_worksheets.cell(row=2, column=1)
        src_end_cell = cai888_worksheets.cell(row=cai888_worksheets.max_row, column=4)
        for i in range(self.external_hotrunner_num):
            dst_start_cell = hotrunner_io_worksheet.cell(row=cur_work_row, column=cur_work_column)
            if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell) != 0:
                print("Fatal: 单元格区域复制失败")
                return -2
            hotrunner_io_worksheet.cell(row=cur_work_row, column=cur_work_column).value += str(i)
            cur_work_row += cai888_worksheets.max_row - 1
        return 0

    def copyEnergyModuleInfo(self):
        ''' 往 硬件配置 中复制DEE模块信息(DEE模块的 硬件排布 修改在modifyImmInfo方法中) '''
        if not self.energy_dee:
            return 0
        module_content_worksheet = self.modulelib_workbook['硬件检索']
        hardware_info_worksheet = self.std_workbook['硬件配置']
        # 定位“特殊功能模块”单元格
        result = searchCells(hardware_info_worksheet, '特殊功能模块')
        if result is None:
            print("Fatal: 在%s中匹配关键字: '特殊功能模块' 失败" % self.std_filename)
            return -1
        first_cell = hardware_info_worksheet.cell(row=result.row + 1, column=result.col_idx)
        # 获取开始工作的单元格行位置
        cur_work_column = first_cell.col_idx
        if first_cell.value is None:
            cur_work_row = first_cell.row
        else:
            cur_work_row = first_cell.row
            cur_work_row += 1
            first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
            while first_cell.value is not None:
                cur_work_row += 1
                first_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        # 向 '硬件配置' 表复制DEE021模块信息
        hardware_info_worksheet.insert_rows(cur_work_row)
        src_end_cell = searchCells(module_content_worksheet, 'DEE021')
        if src_end_cell is None:
            print("Fatal: 在 %s->%s 中匹配关键字'DEE021'失败 " % (self.modulelib_filename, module_content_worksheet.title))
            return -2
        src_start_cell = module_content_worksheet.cell(row=src_end_cell.row, column=1)
        dst_start_cell = hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column)
        if self._copyAndModifyCellRange(src_start_cell, src_end_cell, dst_start_cell, style=False) != 0:
            print("Fatal: 单元格区域复制失败")
            return -3
        hardware_info_worksheet.cell(row=cur_work_row, column=cur_work_column + 3, value='+1')

    def saveAs(self, filename):
        self.std_workbook.save(filename)
        return 0



if __name__ == '__main__':
    iofile = IOFile(imm_type='ve2s',
                    main_board_modules=[['cdm163', {'I1': ['点1', 'id 1']}], ['cio011', {'O2': ['点45', 'id 45'], 'Ai2': ['模腔压力', 'Mold Pressure']}], ['cto163', {'O1': ['点41', 'ID 41']}]],
                    board_1_modules=[['civ512', {}], ['cai888', {}]],
                    big=True)
    iofile.copyMainBoardModulesInfo()
    iofile.copyMainBoardModules()
    iofile.copyBoard1ModulesInfo()
    iofile.copyBoard1Modules()
    iofile.saveAs('../1.xlsx')
